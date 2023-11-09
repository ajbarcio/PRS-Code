import numpy
import openpyxl
from openpyxl import *

PATH = 'C:\Stuff\SMM\PRS-optimizzation.xlsx'

wb = load_workbook(filename = PATH)
ws = wb.active

fuck = [1.2, 3.4, 5.6]
you = [ws['C2'].value,ws['C3'].value,ws['C4'].value]

print(f"{you[0]},{you[1]},{you[2]}")